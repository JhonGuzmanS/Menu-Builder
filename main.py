import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_folder(item, options, suboptions):
    with open('testings.csv', 'a', newline='') as csvfile:
        fieldnames = ['Menu Item Full Name', 'Menu Item Group', 'Menu Item Category', 'Default Price', 'Dine In Price', 'Bar Price', 'Pick Up Price', 'Take Out Price', 'Delivery Price', 'Open Price Item', 'POS Orders Print At', 'Tax 1', 'Tax 2', 'Tax 3', 'This Is A Bar Item', 'This Is A Weighted Item', 'Tare', 'Barcode', 'Item Folder', 'Belongs To Item Folder']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
    
            # For the folder menu items
        for option in options:
            folder_name = f"{option} {item}"
            print(f"{folder_name}")

                # For the subfolders menu items
            for sub in suboptions:
                subfolder_name = f"{folder_name} - {sub}"
                print(f"{subfolder_name}")
                writer.writerow({'Menu Item Full Name': subfolder_name, 'Menu Item Group': 'Coffee', 'Menu Item Category': 'Drinks', 'Default Price': '1', 'POS Orders Print At': 'N', 'Item Folder': '1', 'Belongs To Item Folder': folder_name})                


def read_csv():
    with open('testings.csv', mode='r') as csvfile:
        return csv.DictReader(csvfile)
        for row in reader:
            print(row)
            #print(f"{row['Menu Item Full Name']} | {row['Menu Item Group']} | {row['Menu Item Category']} | $ {row['Default Price']} ")


def conv_pd():
    df = pd.read_csv('testings.csv')
    print(df)
    return df


def create_template_file(headers: list[str], bold_headers: list[str], filepath: str = "Template.xlsx") -> None:
    """
    Creates a new .xlsx file with a formatted header row.

    :param headers: list of header names, in column order (must be <= 20 items)
    :param bold_headers: list of header names (subset of `headers`) that should be bolded
    :param filepath: path to save the new .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    num_columns = 20
    bold_set = set(bold_headers)

    for col_idx in range(1, num_columns + 1):
        header_text = headers[col_idx - 1] if col_idx <= len(headers) else ""
        header_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = Font(name="Calibri", size=11, bold=header_text in bold_set)
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
        cell.fill = header_fill

    thin = Side(style="thin")
    for col_idx in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col_idx)
        left = thin if col_idx == 1 else None
        right = thin if col_idx == num_columns else None
        cell.border = Border(top=thin, bottom=thin, left=left, right=right)

    wb.save(filepath)