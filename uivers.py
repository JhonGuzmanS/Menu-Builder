from nicegui import ui
from openpyxl import load_workbook
from io import BytesIO
from main import create_template_file
from MenuItem import MenuItem
from datetime import datetime


#------------------------Start of Functions------------------------------------------------

def add_row(fn, it, ig, ic, ip):
    #ui.label(f"Added: {fn.value + '-' + it.value} | {ip.value} | {fn.value} | {ig.value} | {ic.value}").update()
    if len(fn.value) > 0:
        table_l.add_item(fn.value + ' ' + it.value, ig.value, ic.value, ip.value, fn.value, print_loc='N')
    else:
        table_l.add_item(it.value, ig.value, ic.value, ip.value, fn.value, print_loc='N')
    

# include limit of 32 per unique group / POS default menu setup
def add(grid,fn, it, ig, ic, ip):
    add_row(fn, it, ig, ic, ip)
    row = table_l.holder[-1]
    with grid.props.suspend_updates():
        grid.options['rowData'].append(row)
    grid.run_grid_method('applyTransaction', {'add': [row]})
    grid.run_grid_method('ensureIndexVisible', len(grid.options['rowData']) - 1)


def add_ids(data):
    for i, row in enumerate(data):
        row['id'] = i
    return data

# create a new excel file, have option to rename/upload to downloads
async def export(grid, filename = f"Menu_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"):
    create_template_file(headerName, ['Menu Item Full Name', 'Menu Item Group', 'Menu Item Category', 'Default Price', 'POS Order Print At'], filename)
    data = await show_grid(grid)
    wb = load_workbook(filename)
    ws = wb["Sheet1"]

    headers = [cell.value for cell in ws[1]]
    for row in data:
        ws.append([row.get(h, None) for h in headers])
    wb.save(filename)
    

async def import_data(e, grid):
    file_bytes = await e.file.read()
    data = import_xlsx(file_bytes)
    data = add_ids(data)
    grid.update()
    grid.options["rowData"] = data
    table_l.update_index(len(data))


def import_xlsx(file_bytes, sheet_name=None):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    
    return [dict(zip(headers, row)) for row in rows]


async def show_items(grid):
    row = await grid.get_client_data()
    print(row)

async def show_grid(grid):
    return await grid.get_client_data()
    print(row)

async def output_folder(folder, grid):
    rows = await grid.get_selected_rows()
    
    if rows:
        for row in rows:
            await grid.run_row_method(row['id'], 'setDataValue', 'Belongs To Item Folder', folder.value, )
            ui.notify(f"Selected row: {row['Menu Item Full Name']}")
            
    else:
        ui.notify('No rows selected.')
    await grid.run_grid_method('deselectAllFiltered')
    await grid.run_grid_method('ensureIndexVisible', rows[0]['id'], 'top')

  
'''
    
    if rows:
        for row in rows:
            index = table_l.get_index(row['Menu Item Full Name'])
            if index != -1:
                grid.options["rowData"][index]['Belongs To Item Folder'] = folder.value

     '''
#------------------------End of Functions------------------------------------------------



# Global variables go here
table_l = MenuItem()    # MenuItem may not be needed  / holds unique ID
printer_options = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'B', 'N']

columns = [
    {'field': 'Menu Item Full Name'},
    {'field': 'Menu Item Group'},
    {'field': 'Menu Item Category'},
    {'field': 'Default Price', 'cellEditor': 'agNumberCellEditor','cellEditorParams': {'min': 0, 'max': 10000, 'precision': 2}},
    {'field': 'Dine In Price', 'hide': True, 'cellEditor': 'agNumberCellEditor','cellEditorParams': {'min': 0, 'max': 10000, 'precision': 2}},
    {'field': 'Bar Price', 'hide': True, 'cellEditor': 'agNumberCellEditor','cellEditorParams': {'min': 0, 'max': 10000, 'precision': 2}},
    {'field': 'Pick Up Price', 'hide': True, 'cellEditor': 'agNumberCellEditor','cellEditorParams': {'min': 0, 'max': 10000, 'precision': 2}},
    {'field': 'Take Out Price', 'hide': True, 'cellEditor': 'agNumberCellEditor','cellEditorParams': {'min': 0, 'max': 10000, 'precision': 2}},
    {'field': 'Delivery Price', 'hide': True, 'cellEditor': 'agNumberCellEditor','cellEditorParams': {'min': 0, 'max': 10000, 'precision': 2}},
    {'field': 'Open Price Item', 'hide': True, 'cellRenderer': 'agCheckboxCellRenderer', 'cellEditor': 'agCheckboxCellEditor',},
    {'field': 'POS Orders Print At', 'cellEditor': 'agSelectCellEditor', 'cellEditorParams':{'values': printer_options}},
    {'field': 'Tax 1', 'cellRenderer': 'agCheckboxCellRenderer', 'cellEditor': 'agCheckboxCellEditor',},
    {'field': 'Tax 2', 'hide': True, 'cellRenderer': 'agCheckboxCellRenderer', 'cellEditor': 'agCheckboxCellEditor',},
    {'field': 'Tax 3', 'hide': True, 'cellRenderer': 'agCheckboxCellRenderer', 'cellEditor': 'agCheckboxCellEditor',},
    {'field': 'This Is A Bar Item', 'hide': True, 'cellRenderer': 'agCheckboxCellRenderer', 'cellEditor': 'agCheckboxCellEditor',},
    {'field': 'This Is A Weighted Item', 'hide': True, 'cellRenderer': 'agCheckboxCellRenderer', 'cellEditor': 'agCheckboxCellEditor',},
    {'field': 'Tare', 'hide': True},
    {'field': 'Barcode', 'hide': True},
    {'field': 'Item Folder', 'cellRenderer': 'agCheckboxCellRenderer', 'cellEditor': 'agCheckboxCellEditor',},
    {'field': 'Belongs To Item Folder',},
]

headerName = [x['field'] for x in columns]
hiddenHeaders = ['Dine In Price','Bar Price','Pick Up Price','Take Out Price','Delivery Price','Open Price Item','Tax 2','Tax 3','This Is A Bar Item','This Is A Weighted Item','Tare','Barcode']
rows = []





#------------------------Start of Page------------------------------------------------

def run():
    with ui.grid(columns=5): # Menu Item Input
        it = ui.input(label="Item Name", value="SM", validation={'field required': lambda val: len(val) > 0})

        ip = ui.number(label="Price", value = 1, validation={'field required': lambda val: val > -1})

        fn = ui.input(label='Belongs to Item Folder')

        ig = ui.input(label='Item Group', value="Coffee", validation={'field required': lambda val: len(val) > 0})

        ic = ui.input(label='Item Category', value="Drinks", validation={'field required': lambda val: len(val) > 0})



    ui.button('Add row', icon='add', on_click=lambda: add(grid, fn, it, ig, ic, ip))
    folder = ui.input(label='Folder Name', value="Hot Coffee")
    ui.button('Add folder', icon='add', on_click=lambda: output_folder(folder, grid))

    grid = ui.aggrid({
        'columnDefs': columns,
        'defaultColDef': {'wrapHeaderText': True,'autoHeaderHeight': True, 'editable': True, 'enableCellChangeFlash': True},
        'rowData': rows,
        'stopEditingWhenCellsLoseFocus': True,
        'rowSelection': {'mode': 'multiRow'},
    })

    ui.upload(on_upload=lambda e: import_data(e, grid), auto_upload=True)
    ui.button('Terminal print', on_click=lambda: show_items(grid))
    ui.button('Show hidden', on_click=lambda: grid.run_grid_method('setColumnsVisible', hiddenHeaders, True))
    ui.button('Hide hidden', on_click=lambda: grid.run_grid_method('setColumnsVisible', hiddenHeaders, False))
    ui.button('Export to Excel', on_click=lambda: export(grid))


    
    ui.run()

run()